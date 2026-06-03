from config import (
    OUTPUT_EXCEL_FILE,
    INSERT_OUTPUT_EXCEL_FILE,
    MONTHLY_SCHEDULE_SHEET_SUFFIX,
    ORDER_VIEW_SHEET_NAME,
    MACHINE_ALLOCATION_SHEET_NAME,
    DATE_MACHINE_SUMMARY_SHEET_NAME,
)


def export_to_excel(
    order_df,
    calendar_df,
    detail_df,
    machine_df=None,
    date_machine_df=None,
    output_file=None,
):
    """
    普通排产导出。

    Sheet1：表1_订单视图
    Sheet2及以后：按月份自动生成排产图
    倒数第二个 Sheet：表3_工序机台数明细 (按订单, 如果 machine_df 非空)
    最后一个 Sheet：表4_按日期机台数汇总 (按日期 + M/A 对比, 如果 date_machine_df 非空)
    """
    if output_file is None:
        output_file = OUTPUT_EXCEL_FILE

    return export_monthly_schedule_to_excel(
        order_df=order_df,
        calendar_df=calendar_df,
        detail_df=detail_df,
        machine_df=machine_df,
        date_machine_df=date_machine_df,
        output_file=output_file,
    )


def export_insert_to_excel(
    order_df,
    new_calendar_df,
    new_detail_df,
    machine_df=None,
    date_machine_df=None,
    previous_plan_file=None,
    output_file=None,
):
    """
    插单模式导出。

    Sheet1：表1_订单视图
    Sheet2及以后：按月份自动生成插单后的新排产图
    倒数第二个 Sheet：表3_工序机台数明细 (按订单)
    最后一个 Sheet：表4_按日期机台数汇总 (按日期 + M/A 对比)

    说明：
    previous_plan_file 参数保留, 是为了兼容旧 main.py 中原来的调用方式。
    新版导出不再把旧计划复制到插单结果文件中。
    """
    if output_file is None:
        output_file = INSERT_OUTPUT_EXCEL_FILE

    return export_monthly_schedule_to_excel(
        order_df=order_df,
        calendar_df=new_calendar_df,
        detail_df=new_detail_df,
        machine_df=machine_df,
        date_machine_df=date_machine_df,
        output_file=output_file,
    )


def export_monthly_schedule_to_excel(
    order_df,
    calendar_df,
    detail_df,
    output_file,
    machine_df=None,
    date_machine_df=None,
):
    """
    按月份导出排产结果。
    """
    month_sheets = _split_calendar_and_detail_by_month(
        calendar_df=calendar_df,
        detail_df=detail_df,
    )

    with __import__("pandas").ExcelWriter(output_file, engine="openpyxl") as writer:
        order_df.to_excel(
            writer,
            sheet_name=ORDER_VIEW_SHEET_NAME,
            index=False,
            startrow=1
        )

        workbook = writer.book

        ws1 = workbook[ORDER_VIEW_SHEET_NAME]
        _format_order_sheet(ws1)

        for item in month_sheets:
            sheet_name = item["sheet_name"]
            calendar_month_df = item["calendar_df"]
            detail_month_df = item["detail_df"]

            calendar_startrow = 1
            detail_gap_rows = 2

            calendar_header_excel_row = calendar_startrow + 1
            calendar_last_data_excel_row = calendar_header_excel_row + len(calendar_month_df)

            detail_title_excel_row = calendar_last_data_excel_row + detail_gap_rows
            detail_startrow = detail_title_excel_row

            calendar_month_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=calendar_startrow
            )

            detail_month_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=detail_startrow
            )

            ws = workbook[sheet_name]

            _format_calendar_and_detail_sheet(
                ws,
                calendar_df=calendar_month_df,
                detail_df=detail_month_df,
                detail_title_excel_row=detail_title_excel_row,
                title_text=sheet_name,
                detail_title_text=f"{sheet_name}-订单日产量明细",
            )

        # =========================
        # 表3_工序机台数明细 (按订单)
        # =========================
        if machine_df is not None and len(machine_df) > 0:
            machine_df.to_excel(
                writer,
                sheet_name=MACHINE_ALLOCATION_SHEET_NAME,
                index=False,
                startrow=1,
            )

            ws_machine = workbook[MACHINE_ALLOCATION_SHEET_NAME]
            _format_machine_allocation_sheet(
                ws_machine,
                machine_df=machine_df,
            )

        # =========================
        # 表4_按日期机台数汇总 (按日期 + M/A 对比)
        # =========================
        if date_machine_df is not None and len(date_machine_df) > 0:
            date_machine_df.to_excel(
                writer,
                sheet_name=DATE_MACHINE_SUMMARY_SHEET_NAME,
                index=False,
                startrow=1,
            )

            ws_date_machine = workbook[DATE_MACHINE_SUMMARY_SHEET_NAME]
            _format_date_machine_summary_sheet(
                ws_date_machine,
                date_machine_df=date_machine_df,
            )

    return output_file


def _is_date_column(column_name):
    """
    判断某一列是否是日期列。

    当前 result_parser.py 生成的日期列格式为：
        5/1
        5/2
        6/1
        12/31

    因此这里用简单规则判断：
        月/日
    """
    import re

    text = str(column_name).strip()

    return re.match(r"^\d{1,2}/\d{1,2}$", text) is not None


def _get_month_from_date_column(column_name):
    """
    从日期列名中提取月份。

    示例：
        5/1 -> 5
        6/30 -> 6
    """
    text = str(column_name).strip()
    month_text = text.split("/", 1)[0]
    return int(month_text)


def _split_calendar_and_detail_by_month(calendar_df, detail_df):
    """
    按月份拆分产线日历和订单日产量明细。

    输入：
        calendar_df：
            第一列是“产线”，后面是日期列，例如 5/1、5/2、6/1。

        detail_df：
            第一列是“订单”，后面是日期列，例如 5/1、5/2、6/1。

    输出：
        [
            {
                "month": 5,
                "sheet_name": "5月排产图",
                "calendar_df": 5月产线日历,
                "detail_df": 5月订单日产量明细,
            },
            {
                "month": 6,
                "sheet_name": "6月排产图",
                "calendar_df": 6月产线日历,
                "detail_df": 6月订单日产量明细,
            },
        ]
    """

    if "产线" not in calendar_df.columns:
        raise ValueError("calendar_df 中缺少 '产线' 列，无法按月份导出排产图。")

    if "订单" not in detail_df.columns:
        raise ValueError("detail_df 中缺少 '订单' 列，无法按月份导出订单日产量明细。")

    month_to_columns = {}

    for col in calendar_df.columns:
        if col == "产线":
            continue

        if not _is_date_column(col):
            continue

        month = _get_month_from_date_column(col)

        if month not in month_to_columns:
            month_to_columns[month] = []

        month_to_columns[month].append(col)

    if not month_to_columns:
        raise ValueError("没有从排产结果中识别到日期列，无法生成月份排产图。")

    month_items = []

    for month in month_to_columns:
        cols = month_to_columns[month]

        calendar_month_df = calendar_df[["产线"] + cols].copy()

        # detail_df 理论上与 calendar_df 具有同样的日期列；
        # 这里做一次保护，避免某些日期列不存在时报错。
        detail_cols = [
            col for col in cols
            if col in detail_df.columns
        ]

        detail_month_df = detail_df[["订单"] + detail_cols].copy()

        sheet_name = f"{month}{MONTHLY_SCHEDULE_SHEET_SUFFIX}"

        month_items.append({
            "month": month,
            "sheet_name": sheet_name,
            "calendar_df": calendar_month_df,
            "detail_df": detail_month_df,
        })

    return month_items


def _copy_sheet_from_existing_workbook(
    source_file,
    source_sheet_name,
    target_file,
    target_sheet_name,
    insert_index=1,
):
    """
    将旧结果文件中的某个 sheet 复制到新结果文件中。

    插单模式下用于：
    - 从 CP_SAT_排产结果.xlsx 复制原始“表2_产线日历”；
    - 放入 CP_SAT_插单排产结果.xlsx；
    - 保证表2是原计划，表3是插单后新计划。

    说明：
    该函数保留用于兼容旧版导出逻辑。
    新版导出已经改为“按月份输出排产图”，默认不再调用该函数。
    """
    from copy import copy
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    source_wb = load_workbook(source_file)
    target_wb = load_workbook(target_file)

    if source_sheet_name not in source_wb.sheetnames:
        raise ValueError(
            f"旧结果文件 {source_file} 中没有找到 sheet：{source_sheet_name}"
        )

    if target_sheet_name in target_wb.sheetnames:
        del target_wb[target_sheet_name]

    source_ws = source_wb[source_sheet_name]

    target_ws = target_wb.create_sheet(
        title=target_sheet_name,
        index=insert_index,
    )

    # =========================
    # 复制单元格内容和样式
    # =========================
    for row in source_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            new_cell = target_ws.cell(
                row=cell.row,
                column=cell.column,
                value=cell.value,
            )

            if cell.has_style:
                new_cell._style = copy(cell._style)

            if cell.number_format:
                new_cell.number_format = cell.number_format

            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)

            if cell.font:
                new_cell.font = copy(cell.font)

            if cell.fill:
                new_cell.fill = copy(cell.fill)

            if cell.border:
                new_cell.border = copy(cell.border)

    # =========================
    # 复制合并单元格
    # =========================
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # =========================
    # 复制列宽
    # =========================
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width

    # =========================
    # 复制行高
    # =========================
    for row_idx, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = row_dim.height

    # =========================
    # 复制冻结窗格和筛选
    # =========================
    target_ws.freeze_panes = source_ws.freeze_panes

    if source_ws.auto_filter and source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref

    target_wb.save(target_file)


def _get_base_styles():
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="F4F9FD")
    header_font = Font(bold=True, color="000000")
    title_font = Font(size=14, bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return {
        "header_fill": header_fill,
        "title_fill": title_fill,
        "header_font": header_font,
        "title_font": title_font,
        "border": border,
        "center": center,
    }


def _format_header_row(ws, row_idx, max_col):
    styles = _get_base_styles()

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]


def _format_range(ws, start_row, end_row, start_col, end_col):
    styles = _get_base_styles()

    if end_row < start_row:
        return

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = styles["border"]
            cell.alignment = styles["center"]


def _format_order_sheet(ws):
    from openpyxl.utils import get_column_letter

    styles = _get_base_styles()

    ws["A1"] = "表1：订单视图"
    ws["A1"].font = styles["title_font"]
    ws["A1"].fill = styles["title_fill"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    _format_header_row(ws, 2, ws.max_column)
    _format_range(ws, 3, ws.max_row, 1, ws.max_column)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    widths = {
        "订单": 18,
        "内部订单名": 24,
        "插单处理方式": 14,
        "是否插单影响": 14,
        "是否插单": 10,
        "是否加量": 10,
        "原需求量": 14,
        "插单增加量": 14,
        "最终需求量": 14,
        "所需产线天数": 14,
        "剩余窗口天数": 14,
        "自动紧迫度": 14,
        "紧迫度权重": 14,
        "窗口开始": 12,
        "窗口结束": 12,
        "实际开工": 12,
        "实际完工": 12,
        "生产天数": 12,
        "总产线·天数": 14,
        "需求量": 14,
        "实际产量": 14,
        "超产线天数": 14,
        "超产量": 14,
        "是否延期": 10,
        "延期天数": 12,
        "延迟天数": 12,
    }

    # 根据表头名称设置列宽。
    # 这样即使 result_parser.py 后续调整字段顺序，也不会导致列宽错位。
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=2, column=col_idx).value
        header_text = str(header_value).strip() if header_value is not None else ""
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(header_text, 14)

    number_headers = {
        "原需求量",
        "插单增加量",
        "最终需求量",
        "所需产线天数",
        "剩余窗口天数",
        "紧迫度权重",
        "生产天数",
        "总产线·天数",
        "需求量",
        "实际产量",
        "超产线天数",
        "超产量",
        "延期天数",
        "延迟天数",
    }

    decimal_headers = {
        "自动紧迫度",
    }

    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=2, column=col_idx).value
        header_text = str(header_value).strip() if header_value is not None else ""

        if header_text in number_headers:
            for cell in ws.iter_cols(
                min_col=col_idx,
                max_col=col_idx,
                min_row=3,
                max_row=ws.max_row,
            ):
                for item in cell:
                    item.number_format = "#,##0"

        if header_text in decimal_headers:
            for cell in ws.iter_cols(
                min_col=col_idx,
                max_col=col_idx,
                min_row=3,
                max_row=ws.max_row,
            ):
                for item in cell:
                    item.number_format = "0.0000"


def _format_calendar_and_detail_sheet(
    ws,
    calendar_df,
    detail_df,
    detail_title_excel_row,
    title_text="表2：产线日历",
    detail_title_text="表2-附表：订单日产量明细",
):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    styles = _get_base_styles()

    # =========================
    # 上方标题：产线日历
    # =========================
    ws["A1"] = title_text
    ws["A1"].font = styles["title_font"]
    ws["A1"].fill = styles["title_fill"]
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=calendar_df.shape[1]
    )

    calendar_header_row = 2
    calendar_first_data_row = 3
    calendar_last_data_row = calendar_first_data_row + len(calendar_df) - 1

    _format_header_row(ws, calendar_header_row, calendar_df.shape[1])
    _format_range(ws, calendar_first_data_row, calendar_last_data_row, 1, calendar_df.shape[1])

    # =========================
    # 下方标题：订单日产量明细
    # =========================
    ws.cell(row=detail_title_excel_row, column=1).value = detail_title_text
    ws.cell(row=detail_title_excel_row, column=1).font = styles["title_font"]
    ws.cell(row=detail_title_excel_row, column=1).fill = styles["title_fill"]
    ws.merge_cells(
        start_row=detail_title_excel_row,
        start_column=1,
        end_row=detail_title_excel_row,
        end_column=detail_df.shape[1]
    )

    detail_header_row = detail_title_excel_row + 1
    detail_first_data_row = detail_header_row + 1
    detail_last_data_row = detail_first_data_row + len(detail_df) - 1

    _format_header_row(ws, detail_header_row, detail_df.shape[1])
    _format_range(ws, detail_first_data_row, detail_last_data_row, 1, detail_df.shape[1])

    # =========================
    # 自动筛选与冻结
    # =========================
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(calendar_df.shape[1])}{calendar_last_data_row}"

    # =========================
    # 订单颜色池：上下两个表保持一致
    # =========================
    color_pool = [
        "E2F0D9",
        "D9EAF7",
        "FCE4D6",
        "EAD1DC",
        "FFF2CC",
        "D9EAD3",
        "D0E0E3",
        "EDEDED",
        "F4CCCC",
        "D9D2E9",
        "CFE2F3",
        "F9CB9C",
        "D5E8D4",
        "F8CECC",
        "DAE8FC",
        "E1D5E7",
    ]

    order_fills = {}
    color_idx = 0

    # 从上方产线日历里按订单第一次出现顺序分配颜色
    for r in range(calendar_first_data_row, calendar_last_data_row + 1):
        for c in range(2, calendar_df.shape[1] + 1):
            value = ws.cell(r, c).value
            if value is not None:
                value = str(value).strip()
                if value != "" and value not in order_fills:
                    order_fills[value] = color_pool[color_idx % len(color_pool)]
                    color_idx += 1

    # =========================
    # 上方产线日历着色
    # =========================
    for r in range(calendar_first_data_row, calendar_last_data_row + 1):
        ws.cell(r, 1).font = Font(bold=True)

        for c in range(2, calendar_df.shape[1] + 1):
            cell = ws.cell(r, c)
            value = cell.value

            if value is not None:
                value = str(value).strip()
                if value in order_fills:
                    cell.fill = PatternFill("solid", fgColor=order_fills[value])

    # =========================
    # 下方订单日产量明细着色
    # 下方表结构：A=订单，B起=日期
    # =========================
    if len(detail_df) > 0:
        summary_fill = PatternFill("solid", fgColor="D9EAF7")
        summary_font = Font(bold=True)

        for r in range(detail_first_data_row, detail_last_data_row + 1):
            order_cell = ws.cell(r, 1)
            order_cell.font = Font(bold=True)

            order_name = order_cell.value

            if order_name is not None:
                order_name = str(order_name).strip()

                # 汇总行：线体合计、产能合计
                if order_name in ["线体合计", "产能合计"]:
                    for c in range(1, detail_df.shape[1] + 1):
                        cell = ws.cell(r, c)
                        cell.fill = summary_fill
                        cell.font = summary_font

                        # A列是文字，不设置数字格式
                        if c == 1:
                            continue

                        # 显示 18、4,140,000 这类数值
                        cell.number_format = "#,##0"
                    continue

                # 普通订单行按订单颜色处理
                if order_name != "" and order_name in order_fills:
                    fill = PatternFill("solid", fgColor=order_fills[order_name])

                    # 订单列染色
                    order_cell.fill = fill

                    # 有生产量的日期单元格染色
                    for c in range(2, detail_df.shape[1] + 1):
                        cell = ws.cell(r, c)
                        if cell.value not in (None, ""):
                            cell.fill = fill
                            cell.number_format = "#,##0"

    # =========================
    # 列宽
    # =========================
    ws.column_dimensions["A"].width = 14

    # 日期列需要放得下 4,140,000，避免 Excel 显示 #######
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # 行高
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22


def _format_machine_allocation_sheet(ws, machine_df):
    """
    格式化"工序机台数明细"表。

    布局:
        第 1 行: 表标题 (合并单元格)
        第 2 行: 列标题 (订单 / 日期 / 占用产线数 / 工序1-制绒 / ...)
        第 3 行起: 数据行, 同一订单使用一致的浅色背景, 便于按订单区分。

    数据来自 scheduler_results.machine_allocation.parse_machine_allocation_view,
    每行表示一个 (订单, 日期) 的机台数分配。
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    styles = _get_base_styles()

    num_cols = machine_df.shape[1]
    num_rows = machine_df.shape[0]

    # =========================
    # 表标题
    # =========================
    ws["A1"] = "表3：工序机台数明细 (基于产线数 l[j,t] 后处理计算)"
    ws["A1"].font = styles["title_font"]
    ws["A1"].fill = styles["title_fill"]
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=num_cols,
    )

    header_row = 2
    first_data_row = 3
    last_data_row = first_data_row + num_rows - 1

    _format_header_row(ws, header_row, num_cols)
    _format_range(ws, first_data_row, last_data_row, 1, num_cols)

    # =========================
    # 按订单分配颜色, 便于阅读
    # =========================
    color_pool = [
        "E2F0D9",
        "D9EAF7",
        "FCE4D6",
        "EAD1DC",
        "FFF2CC",
        "D9EAD3",
        "D0E0E3",
        "EDEDED",
        "F4CCCC",
        "D9D2E9",
        "CFE2F3",
        "F9CB9C",
        "D5E8D4",
        "F8CECC",
        "DAE8FC",
        "E1D5E7",
    ]

    order_fills = {}
    color_idx = 0

    for r in range(first_data_row, last_data_row + 1):
        order_value = ws.cell(r, 1).value
        if order_value is None:
            continue

        order_name = str(order_value).strip()
        if order_name == "":
            continue

        if order_name not in order_fills:
            order_fills[order_name] = color_pool[color_idx % len(color_pool)]
            color_idx += 1

    for r in range(first_data_row, last_data_row + 1):
        order_value = ws.cell(r, 1).value
        if order_value is None:
            continue

        order_name = str(order_value).strip()
        if order_name == "" or order_name not in order_fills:
            continue

        fill = PatternFill("solid", fgColor=order_fills[order_name])
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = fill

    # =========================
    # 列宽
    # =========================
    column_widths = {
        "订单": 18,
        "日期": 10,
        "占用产线数": 12,
    }

    for col_idx in range(1, num_cols + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        header_text = str(header_value).strip() if header_value is not None else ""
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(header_text, 14)

    # =========================
    # 数字格式
    # =========================
    # zgy: 占用产线数 (col 3) 保持整数;
    #      工序机台数列 (col 4+) 按值类型自适应:
    #        - 整数值 (机台数够): "0"
    #        - 浮点值 (机台数不够, 突破 MATRIX_A 上限): "0.00"
    for r in range(first_data_row, last_data_row + 1):
        ws.cell(r, 3).number_format = "0"
        for col_idx in range(4, num_cols + 1):
            cell = ws.cell(r, col_idx)
            val = cell.value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                if float(val).is_integer():
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"

    # =========================
    # 冻结窗格和筛选
    # =========================
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{last_data_row}"


def _format_date_machine_summary_sheet(ws, date_machine_df):
    """
    格式化"按日期机台数汇总"表 (表4)。

    布局:
        第 1 行: 表标题 (合并单元格)
        第 2 行: 列标题 (日期 / 类型/订单 / 占用产线 / 工序1~工序12 / 备注)
        第 3 行起: 数据行, 每个日期一个分块, 包含:
            - 当天每个订单一行 (浅色背景);
            - M矩阵汇总 (深蓝色背景, 加粗);
            - A矩阵理论 (浅黄背景);
            - 差异(M-A) (按正/负配色: 正=红=机台不够, 负=绿=空余);
            - 空行分隔。
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    styles = _get_base_styles()

    num_cols = date_machine_df.shape[1]
    num_rows = date_machine_df.shape[0]

    # 表标题
    ws["A1"] = "表4：按日期机台数汇总 (M矩阵 vs A矩阵理论 对比, 用于切线决策)"
    ws["A1"].font = styles["title_font"]
    ws["A1"].fill = styles["title_fill"]
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=num_cols,
    )

    header_row = 2
    first_data_row = 3
    last_data_row = first_data_row + num_rows - 1

    _format_header_row(ws, header_row, num_cols)
    _format_range(ws, first_data_row, last_data_row, 1, num_cols)

    # =========================
    # 颜色定义
    # =========================
    m_sum_fill = PatternFill("solid", fgColor="2F75B5")        # 深蓝
    m_sum_font = Font(bold=True, color="FFFFFF")               # 白字
    a_ref_fill = PatternFill("solid", fgColor="FFE699")        # 浅黄
    a_ref_font = Font(italic=True)
    diff_pos_fill = PatternFill("solid", fgColor="F8CBAD")     # 红 (机台不够)
    diff_neg_fill = PatternFill("solid", fgColor="C6E0B4")     # 绿 (空余)
    diff_zero_fill = PatternFill("solid", fgColor="D9D9D9")    # 灰 (平衡)
    diff_font_bold = Font(bold=True)

    order_color_pool = [
        "E2F0D9", "FCE4D6", "EAD1DC", "FFF2CC",
        "D9EAD3", "D0E0E3", "F4CCCC", "D9D2E9",
        "CFE2F3", "F9CB9C", "D5E8D4", "F8CECC",
        "DAE8FC", "E1D5E7",
    ]

    order_fills = {}
    color_idx = 0

    # 找到"类型/订单"列 (一般在第 2 列, 但稳妥起见按表头定位)
    type_col_idx = None
    for c in range(1, num_cols + 1):
        header_val = ws.cell(row=header_row, column=c).value
        if header_val is not None and str(header_val).strip() == "类型/订单":
            type_col_idx = c
            break
    if type_col_idx is None:
        type_col_idx = 2

    # 工序列范围: 跳过 日期 / 类型/订单 / 占用产线, 不包括最后的 备注 列
    process_col_start = 4
    process_col_end = num_cols - 1   # 备注是最后一列

    # =========================
    # 逐行配色
    # =========================
    for r in range(first_data_row, last_data_row + 1):
        type_value = ws.cell(r, type_col_idx).value
        if type_value is None:
            continue

        type_text = str(type_value).strip()

        if type_text == "":
            continue

        if type_text == "M矩阵汇总":
            for c in range(1, num_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = m_sum_fill
                cell.font = m_sum_font
            continue

        if type_text.startswith("A矩阵理论"):
            for c in range(1, num_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = a_ref_fill
                cell.font = a_ref_font
            continue

        if type_text == "差异(M-A)":
            # 工序列按数值正/负着色
            for c in range(1, num_cols + 1):
                cell = ws.cell(r, c)
                cell.font = diff_font_bold

            for c in range(process_col_start, process_col_end + 1):
                cell = ws.cell(r, c)
                val = cell.value
                if isinstance(val, (int, float)):
                    if val > 0:
                        cell.fill = diff_pos_fill
                    elif val < 0:
                        cell.fill = diff_neg_fill
                    else:
                        cell.fill = diff_zero_fill
            continue

        # 否则视为订单行: 按订单名分配颜色
        if type_text not in order_fills:
            order_fills[type_text] = order_color_pool[color_idx % len(order_color_pool)]
            color_idx += 1

        fill = PatternFill("solid", fgColor=order_fills[type_text])
        # 订单/类型列加粗
        ws.cell(r, type_col_idx).font = Font(bold=True)
        ws.cell(r, type_col_idx).fill = fill

    # =========================
    # 列宽
    # =========================
    column_widths = {
        "日期": 8,
        "类型/订单": 22,
        "占用产线": 10,
        "备注": 50,
    }

    for col_idx in range(1, num_cols + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        header_text = str(header_value).strip() if header_value is not None else ""
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(header_text, 12)

    # 工序列数字格式
    # zgy: 按单元格值类型自适应:
    #   - 整数值 (机台数够 / A 矩阵理论): "0"
    #   - 浮点值 (机台数不够 / 差异有小数): "0.00"
    for col_idx in range(process_col_start, process_col_end + 1):
        for r in range(first_data_row, last_data_row + 1):
            cell = ws.cell(r, col_idx)
            val = cell.value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                if float(val).is_integer():
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"

    # 冻结窗格 + 筛选
    ws.freeze_panes = "D3"


def print_monthly_sheet_info(display_dates):
    """
    根据展示日期打印本次会导出的月份排产图名称。

    例如：
    display_dates 覆盖 2026-05-01 ~ 2026-06-30，
    则控制台打印：
        Sheet 2：5月排产图
        Sheet 3：6月排产图
    """
    months = []

    for display_date in display_dates:
        month = display_date.month
        if month not in months:
            months.append(month)

    for idx, month in enumerate(months, start=2):
        print(f"Sheet {idx}：{month}月排产图")