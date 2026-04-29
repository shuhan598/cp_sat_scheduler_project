from config import OUTPUT_EXCEL_FILE


def export_to_excel(order_df, calendar_df, detail_df, output_file=None):
    if output_file is None:
        output_file = OUTPUT_EXCEL_FILE

    calendar_startrow = 1
    detail_gap_rows = 2

    calendar_header_excel_row = calendar_startrow + 1
    calendar_last_data_excel_row = calendar_header_excel_row + len(calendar_df)

    detail_title_excel_row = calendar_last_data_excel_row + detail_gap_rows
    detail_startrow = detail_title_excel_row

    with __import__("pandas").ExcelWriter(output_file, engine="openpyxl") as writer:
        order_df.to_excel(
            writer,
            sheet_name="表1_订单视图",
            index=False,
            startrow=1
        )

        calendar_df.to_excel(
            writer,
            sheet_name="表2_产线日历",
            index=False,
            startrow=calendar_startrow
        )

        detail_df.to_excel(
            writer,
            sheet_name="表2_产线日历",
            index=False,
            startrow=detail_startrow
        )

        workbook = writer.book

        ws1 = workbook["表1_订单视图"]
        _format_order_sheet(ws1)

        ws2 = workbook["表2_产线日历"]
        _format_calendar_and_detail_sheet(
            ws2,
            calendar_df=calendar_df,
            detail_df=detail_df,
            detail_title_excel_row=detail_title_excel_row
        )

    return output_file


def _get_base_styles():
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="F4F9FD")
    header_font = Font(bold=True, color="000000")
    title_font = Font(size=14, bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return {
        "header_fill": header_fill,
        "title_fill": title_fill,
        "header_font": header_font,
        "title_font": title_font,
        "border": border,
        "center": center,
    }


def _format_header_row(ws, row_idx, max_col):
    styles = _get_base_styles()

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]


def _format_range(ws, start_row, end_row, start_col, end_col):
    styles = _get_base_styles()

    if end_row < start_row:
        return

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = styles["border"]
            cell.alignment = styles["center"]


def _format_order_sheet(ws):
    from openpyxl.utils import get_column_letter

    styles = _get_base_styles()

    ws["A1"] = "表1：订单视图"
    ws["A1"].font = styles["title_font"]
    ws["A1"].fill = styles["title_fill"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    _format_header_row(ws, 2, ws.max_column)
    _format_range(ws, 3, ws.max_row, 1, ws.max_column)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    widths = {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 12,
    }

    for col, width in widths.items():
        if ws.max_column >= ord(col) - ord("A") + 1:
            ws.column_dimensions[col].width = width

    for col in ["G", "H", "I", "J", "K"]:
        if ws.max_column >= ord(col) - ord("A") + 1:
            for cell in ws[col][2:]:
                cell.number_format = "#,##0"


def _format_calendar_and_detail_sheet(ws, calendar_df, detail_df, detail_title_excel_row):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    styles = _get_base_styles()

    # =========================
    # 上方标题：产线日历
    # =========================
    ws["A1"] = "表2：产线日历"
    ws["A1"].font = styles["title_font"]
    ws["A1"].fill = styles["title_fill"]
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=calendar_df.shape[1]
    )

    calendar_header_row = 2
    calendar_first_data_row = 3
    calendar_last_data_row = calendar_first_data_row + len(calendar_df) - 1

    _format_header_row(ws, calendar_header_row, calendar_df.shape[1])
    _format_range(ws, calendar_first_data_row, calendar_last_data_row, 1, calendar_df.shape[1])

    # =========================
    # 下方标题：订单日产量明细
    # =========================
    ws.cell(row=detail_title_excel_row, column=1).value = "表2-附表：订单日产量明细"
    ws.cell(row=detail_title_excel_row, column=1).font = styles["title_font"]
    ws.cell(row=detail_title_excel_row, column=1).fill = styles["title_fill"]
    ws.merge_cells(
        start_row=detail_title_excel_row,
        start_column=1,
        end_row=detail_title_excel_row,
        end_column=detail_df.shape[1]
    )

    detail_header_row = detail_title_excel_row + 1
    detail_first_data_row = detail_header_row + 1
    detail_last_data_row = detail_first_data_row + len(detail_df) - 1

    _format_header_row(ws, detail_header_row, detail_df.shape[1])
    _format_range(ws, detail_first_data_row, detail_last_data_row, 1, detail_df.shape[1])

    # =========================
    # 自动筛选与冻结
    # =========================
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(calendar_df.shape[1])}{calendar_last_data_row}"

    # =========================
    # 订单颜色池：上下两个表保持一致
    # =========================
    color_pool = [
        "E2F0D9",
        "D9EAF7",
        "FCE4D6",
        "EAD1DC",
        "FFF2CC",
        "D9EAD3",
        "D0E0E3",
        "EDEDED",
        "F4CCCC",
        "D9D2E9",
        "CFE2F3",
        "F9CB9C",
        "D5E8D4",
        "F8CECC",
        "DAE8FC",
        "E1D5E7",
    ]

    order_fills = {}
    color_idx = 0

    # 从上方产线日历里按订单第一次出现顺序分配颜色
    for r in range(calendar_first_data_row, calendar_last_data_row + 1):
        for c in range(2, calendar_df.shape[1] + 1):
            value = ws.cell(r, c).value
            if value is not None:
                value = str(value).strip()
                if value != "" and value not in order_fills:
                    order_fills[value] = color_pool[color_idx % len(color_pool)]
                    color_idx += 1

    # =========================
    # 上方产线日历着色
    # =========================
    for r in range(calendar_first_data_row, calendar_last_data_row + 1):
        ws.cell(r, 1).font = Font(bold=True)

        for c in range(2, calendar_df.shape[1] + 1):
            cell = ws.cell(r, c)
            value = cell.value

            if value is not None:
                value = str(value).strip()
                if value in order_fills:
                    cell.fill = PatternFill("solid", fgColor=order_fills[value])

    # =========================
    # 下方订单日产量明细着色
    # 下方表结构：A=订单，B起=日期
    # =========================
    if len(detail_df) > 0:
        summary_fill = PatternFill("solid", fgColor="D9EAF7")
        summary_font = Font(bold=True)

        for r in range(detail_first_data_row, detail_last_data_row + 1):
            order_cell = ws.cell(r, 1)
            order_cell.font = Font(bold=True)

            order_name = order_cell.value

            if order_name is not None:
                order_name = str(order_name).strip()

                # 汇总行：线体合计、产能合计
                if order_name in ["线体合计", "产能合计"]:
                    for c in range(1, detail_df.shape[1] + 1):
                        cell = ws.cell(r, c)
                        cell.fill = summary_fill
                        cell.font = summary_font

                        # A列是文字，不设置数字格式
                        if c == 1:
                            continue

                        # 显示 18、4,140,000 这类数值
                        cell.number_format = "#,##0"
                    continue

                # 普通订单行按订单颜色处理
                if order_name != "" and order_name in order_fills:
                    fill = PatternFill("solid", fgColor=order_fills[order_name])

                    # 订单列染色
                    order_cell.fill = fill

                    # 有生产量的日期单元格染色
                    for c in range(2, detail_df.shape[1] + 1):
                        cell = ws.cell(r, c)
                        if cell.value not in (None, ""):
                            cell.fill = fill
                            cell.number_format = "#,##0"

    # =========================
    # 列宽
    # =========================
    ws.column_dimensions["A"].width = 14

    # 日期列需要放得下 4,140,000，避免 Excel 显示 #######
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # 行高
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22